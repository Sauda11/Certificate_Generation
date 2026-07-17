import pandas as pd
from openpyxl import load_workbook


EXCEL_FILE = "students/students.xlsx"


def read_students(excel_file=EXCEL_FILE):
    """
    Reads all students from the Excel file.
    Returns a list of dictionaries.
    """

    df = pd.read_excel(excel_file)

    students = []

    for _, row in df.iterrows():

        students.append({

            "student_name": str(row["Student Name"]).strip(),
            "course": str(row["Course"]).strip(),
            "certificate_id": str(row["Certificate ID"]).strip(),
            "issue_date": str(row["Issue Date"]).strip(),
            "duration": str(row["Duration"]).strip(),
            "level": str(row["Level"]).strip(),
            "email": str(row["Email"]).strip(),
            "phone": str(row["Phone"]).strip(),
            "status": str(row["Status"]).strip()

        })

    return students


def update_student_status(certificate_id, status):
    """
    Updates the Status column for a student.
    """

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    headers = {}

    for col in range(1, sheet.max_column + 1):
        headers[sheet.cell(row=1, column=col).value] = col

    id_col = headers["Certificate ID"]
    status_col = headers["Status"]

    for row in range(2, sheet.max_row + 1):

        if str(sheet.cell(row=row, column=id_col).value).strip() == certificate_id:

            sheet.cell(row=row, column=status_col).value = status
            break

    workbook.save(EXCEL_FILE)


if __name__ == "__main__":

    students = read_students()

    for student in students:
        print(student)